import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\SivakarnaSatri\\Desktop\\Working\Data1.xlsx")
sheets = wb.sheetnames
#print(wb.active.title)
sh1 = wb['Login']
data = sh1['B2'].value
print(data)

#option1
print(wb['Login']['A2'].value)

#option2
print(sh1.cell(3,2).value)
print(sh1.cell(3,3).value)

sh2 = wb['Mark']
print(sh2.cell(2,2).value)
print(sh2.cell(3,2).value)

#option3
c=sh2.cell(row=2,column=2)
print(sh2.cell(row=2,column=1).value)

print(wb.get_sheet_by_name('Login').cell(row=4,column=1).value)


