import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\SivakarnaSatri\\Desktop\\Working\Data1.xlsx")
sh1 = wb["Login"]
row = sh1.max_row
column = sh1.max_column

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i, j).value)

sh1.cell(row=5,column=1,value='pytest')
sh1.cell(row=5,column=2,value='UK')
sh1.cell(row=5,column=3,value='98.90')

wb.save("Report.xlsx")

